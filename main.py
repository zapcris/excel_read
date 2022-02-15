from tkinter import *
import tkinter as tk
import tksheet
from tkinter.filedialog import askopenfile
from openpyxl import load_workbook
import xlrd
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import numpy as np
import pandas as pd
import sys






def open_file():
    global open_filename


    file = askopenfile(mode ='r', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')]) # To open the file that you want.
    #' mode='r' ' is to tell the filedialog to read the file
    # 'filetypes=[()]' is to filter the files shown as only Excel files

    wb = load_workbook(filename = file.name) # Load into openpyxl
    wb2 = wb.active

    sheets = wb.sheetnames
    sh1 = wb[sheets[0]]
    open_filename = sh1

     #print(wb.sheetnames)
    # print(sheets)

    row = sh1.max_row
    column = sh1.max_column



    batch_seq = [[] for i in range(column)]

    for i in range(1, column+1):
        for j in range(2, row + 1):
            # print(sh1.cell(i,1).value)
            if sh1.cell(j, i).value != None:
                batch_seq[i-1].append(sh1.cell(j, i).value)

    for i in batch_seq:
        print(i)

def clr_dataframe():
 dframe._clear_item_cache()

#
# root = Tk()
# root.geometry('500x400')
# btn = Button(root, text ='Open the batch sequence excel file', command = open_file)
# btn.pack(side='top')
#
# root.mainloop()

root = Tk()
root.geometry('580x400')

btn = Button(root, text ='Open the batch sequence excel file', command = open_file)
btn.pack(side='top')

btn2 = Button(root, text ='clear data', command = clr_dataframe)
btn2.pack(side='bottom')



# dates = pd.date_range('20210101', periods=8)
# dframe = pd.DataFrame(np.random.randn(8,4),index=dates,columns=list('ABCD'))
dframe = pd.DataFrame

txt = Text(root)
txt.pack()

class PrintToTXT(object):
 def write(self, s):
     txt.insert(END, s)

sys.stdout = PrintToTXT()

print ("The Batch Sequence from the sheet")

print (dframe)



mainloop()